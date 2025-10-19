import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import numbers

def revise_product_prices(input_file='product_price_revision.xlsx', 
                         output_file_base='product_prices_revised'):
    """
    Process Excel file to update prices based on 'new price' column.
    Saves output in both XLSX and CSV formats with proper barcode handling.
    
    Args:
        input_file (str): Path to input Excel file
        output_file_base (str): Base name for output files (without extension)
    
    Returns:
        dict: Statistics about the processing
    """
    
    try:
        # Read the Excel file with Barcode as string to preserve full number
        print(f"Reading Excel file: {input_file}")
        
        # First, read to identify columns
        df_temp = pd.read_excel(input_file, nrows=1)
        
        # Set dtype for Barcode column if it exists
        dtype_dict = {}
        if 'Barcode' in df_temp.columns:
            dtype_dict['Barcode'] = str
        if 'SKU' in df_temp.columns:
            dtype_dict['SKU'] = str
        
        # Read the full file with proper dtypes
        df = pd.read_excel(input_file, dtype=dtype_dict)
        
        # Display initial statistics
        print(f"\nInitial data shape: {df.shape}")
        print(f"Columns found: {df.columns.tolist()}")
        
        # Count rows with new prices before processing
        has_new_price = df['new price'].notna()
        num_updated = has_new_price.sum()
        
        print(f"\nStatistics:")
        print(f"Total products: {len(df)}")
        print(f"Products with new prices: {num_updated}")
        print(f"Products without new prices: {len(df) - num_updated}")
        
        # Apply the logic: Replace Price with new price where new price exists
        df.loc[has_new_price, 'Price'] = df.loc[has_new_price, 'new price']
        
        # Select only the required columns
        required_columns = ['Product Id', 'Product Name', 'Tab', 'Category', 'SKU', 'Barcode', 'Price']
        
        # Check if all required columns exist
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"\nWarning: Missing columns: {missing_columns}")
            # Use only available columns
            required_columns = [col for col in required_columns if col in df.columns]
        
        # Create final dataframe with only required columns
        final_df = df[required_columns].copy()
        
        # Handle NaN values and ensure Barcode is string
        if 'SKU' in final_df.columns:
            final_df['SKU'] = final_df['SKU'].fillna('').astype(str)
            # Remove .0 from SKU if it got converted to float
            final_df['SKU'] = final_df['SKU'].str.replace(r'\.0$', '', regex=True)
        
        if 'Barcode' in final_df.columns:
            # Convert Barcode to string and handle NaN
            final_df['Barcode'] = final_df['Barcode'].fillna('').astype(str)
            # Remove scientific notation if any (e.g., "8.8501E+12" -> "8885008900338")
            # Remove .0 suffix if present
            final_df['Barcode'] = final_df['Barcode'].str.replace(r'\.0$', '', regex=True)
            final_df['Barcode'] = final_df['Barcode'].str.replace('nan', '', regex=False)
        
        # Save to CSV format
        csv_file = f"{output_file_base}.csv"
        print(f"\nSaving CSV to: {csv_file}")
        final_df.to_csv(csv_file, index=False, encoding='utf-8-sig')
        
        # Save to Excel format with proper formatting
        xlsx_file = f"{output_file_base}.xlsx"
        print(f"Saving Excel to: {xlsx_file}")
        
        # Write Excel with special handling for Barcode column
        with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='Products')
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Products']
            
            # Format Barcode column as text to preserve full number
            if 'Barcode' in final_df.columns:
                barcode_col_idx = final_df.columns.get_loc('Barcode') + 1  # +1 because Excel is 1-indexed
                for row in range(2, len(final_df) + 2):  # Start from row 2 (after header)
                    cell = worksheet.cell(row=row, column=barcode_col_idx)
                    cell.number_format = '@'  # @ means text format in Excel
            
            # Also format SKU column as text if it exists
            if 'SKU' in final_df.columns:
                sku_col_idx = final_df.columns.get_loc('SKU') + 1
                for row in range(2, len(final_df) + 2):
                    cell = worksheet.cell(row=row, column=sku_col_idx)
                    cell.number_format = '@'
            
            # Auto-adjust column widths for better readability
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Display sample of updated products
        if num_updated > 0:
            print("\nSample of updated products (first 5):")
            updated_indices = df[has_new_price].index[:5]
            for idx in updated_indices:
                product_name = df.loc[idx, 'Product Name']
                new_price = df.loc[idx, 'new price']
                barcode = final_df.loc[idx, 'Barcode'] if 'Barcode' in final_df.columns else 'N/A'
                print(f"  - {product_name}")
                print(f"    New Price: ${new_price:.2f}")
                print(f"    Barcode: {barcode}")
        
        # Verify barcode format in the saved files
        if 'Barcode' in final_df.columns:
            sample_barcodes = final_df[final_df['Barcode'] != '']['Barcode'].head(3)
            print("\nSample barcodes (verification):")
            for barcode in sample_barcodes:
                print(f"  {barcode}")
        
        print(f"\n✅ Success! Files saved:")
        print(f"  📊 Excel: {xlsx_file}")
        print(f"  📄 CSV: {csv_file}")
        
        return {
            'total_products': len(final_df),
            'products_updated': num_updated,
            'products_unchanged': len(final_df) - num_updated,
            'xlsx_file': xlsx_file,
            'csv_file': csv_file
        }
        
    except FileNotFoundError:
        print(f"❌ Error: File '{input_file}' not found!")
        return None
    except Exception as e:
        print(f"❌ Error processing file: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def main():
    """
    Main function to run the price revision process.
    """
    # You can modify these file paths as needed
    input_file = 'product_price_revision.xlsx'
    output_base = 'product_prices_revised'  # Will create .xlsx and .csv
    
    # Process the file
    results = revise_product_prices(input_file, output_base)
    
    if results:
        print("\n" + "="*50)
        print("PROCESSING COMPLETE")
        print("="*50)
        print(f"Total Products: {results['total_products']}")
        print(f"Prices Updated: {results['products_updated']}")
        print(f"Prices Unchanged: {results['products_unchanged']}")
        print(f"Output Files:")
        print(f"  - Excel: {results['xlsx_file']}")
        print(f"  - CSV: {results['csv_file']}")


# Alternative: Quick function to fix barcode formatting in existing file
def fix_barcode_format(file_path):
    """
    Utility function to fix barcode formatting in an existing Excel/CSV file.
    """
    # Determine file type
    if file_path.endswith('.csv'):
        df = pd.read_csv(file_path, dtype={'Barcode': str, 'SKU': str})
        output_file = file_path.replace('.csv', '_fixed.csv')
        df.to_csv(output_file, index=False, encoding='utf-8-sig')
    else:
        df = pd.read_excel(file_path, dtype={'Barcode': str, 'SKU': str})
        output_file = file_path.replace('.xlsx', '_fixed.xlsx')
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Products')
            worksheet = writer.sheets['Products']
            
            # Format Barcode column as text
            if 'Barcode' in df.columns:
                barcode_col_idx = df.columns.get_loc('Barcode') + 1
                for row in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row, column=barcode_col_idx)
                    cell.number_format = '@'
    
    print(f"Fixed file saved as: {output_file}")
    return output_file


# Simplified version with both output formats
def simple_price_revision(input_file='product_price_revision.xlsx'):
    """
    Simplified version that handles both CSV and Excel output with proper barcode formatting.
    """
    # Read with barcode as string
    df = pd.read_excel(input_file, dtype={'Barcode': str, 'SKU': str})
    
    # Update prices where new price exists
    mask = df['new price'].notna()
    df.loc[mask, 'Price'] = df.loc[mask, 'new price']
    
    # Select required columns
    columns = ['Product Id', 'Product Name', 'Tab', 'Category', 'SKU', 'Barcode', 'Price']
    final_df = df[columns].copy()
    
    # Clean up Barcode and SKU
    final_df['Barcode'] = final_df['Barcode'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True)
    final_df['SKU'] = final_df['SKU'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True)
    
    # Save both formats
    final_df.to_csv('product_prices_revised.csv', index=False, encoding='utf-8-sig')
    final_df.to_excel('product_prices_revised.xlsx', index=False, sheet_name='Products')
    
    print(f"✅ Processed {len(df)} products. Updated {mask.sum()} prices.")
    print(f"📁 Files saved: product_prices_revised.xlsx and product_prices_revised.csv")


if __name__ == "__main__":
    # Run the main function
    main()